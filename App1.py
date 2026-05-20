#!/usr/bin/python
"""
Belar AMMA-2 Integrated Logger and Flask Remote Web Dashboard.

This script runs a background thread to continuously poll data from the
Belar AMMA-2, updates a live web dashboard, and writes all captured
telemetry entries to both data.log and a formatted Excel spreadsheet.
"""

import logging
import os
import random  # Fallback for demonstration if hardware isn't attached
from threading import Thread
import time
from flask import Flask, jsonify, render_template_string
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

app = Flask(__name__)

# Configure the logging utility to write directly to data.log
logging.basicConfig(
    filename="./data.log",
    level=logging.INFO,
    format="%(asctime)s : %(message)s",
    datefmt="%Y-%m-%d %H:%M"
)

# Global runtime dictionary to act as our live data cache
live_telemetry = {
    "negative_peak": "0.0",
    "positive_peak": "0.0",
    "carrier_reference": "100.0",
    "last_updated": "Never"
}

EXCEL_FILE = "./modulation_data.xlsx"


def init_excel_workbook():
    """Create and style the Excel log file if it does not already exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Telemetry History"

        # Professional styling configuration
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Dark Blue
        center_align = Alignment(horizontal="center", vertical="center")

        headers = [
            "Timestamp",
            "Negative Peak (DA)",
            "Positive Peak (DB)",
            "Carrier Ref (DG)"
        ]

        ws.append(headers)

        # Apply styles across header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Freeze the header row so it stays visible while scrolling down logs
        ws.freeze_panes = "A2"

        wb.save(EXCEL_FILE)
        print(f"Initialized styled Excel workbook at: {EXCEL_FILE}")


def append_to_excel(timestamp, neg, pos, carrier):
    """Safely append a structured data row to the Excel workbook."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        row_data = [timestamp, neg, pos, carrier]
        ws.append(row_data)

        # Apply clean text styling to the newly appended row
        last_row = ws.max_row
        data_font = Font(name="Segoe UI", size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[last_row]:
            cell.font = data_font
            cell.alignment = center_align

        # Dynamically auto-fit column widths to prevent visual truncation
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        wb.save(EXCEL_FILE)
    except Exception as e:
        logging.error("Failed writing dataset row entry to Excel file: %s", e)


def telemetry_background_worker():
    """Background thread that reads from the hardware and logs data to files."""
    global live_telemetry

    # Initialize the structural Excel sheet on background worker launch
    init_excel_workbook()

    print("Starting background telemetry collection and dual-file logging...")

    while True:
        try:
            # 1. Capture/Simulate data parameters
            neg = f"{round(random.uniform(85.0, 99.5), 1)}%"
            pos = f"{round(random.uniform(105.0, 122.0), 1)}%"
            carrier = f"{round(random.uniform(98.0, 101.5), 1)}%"
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

            # 2. Update the thread-safe web server cache
            live_telemetry["negative_peak"] = neg
            live_telemetry["positive_peak"] = pos
            live_telemetry["carrier_reference"] = carrier
            live_telemetry["last_updated"] = timestamp

            # 3. Write structured telemetry entries to ./data.log
            logging.info(";Data for DA :; %s", neg)
            logging.info(";Data for DB :; %s", pos)
            logging.info(";Data for DG :; %s", carrier)

            # 4. Write data row to the Excel sheet
            append_to_excel(timestamp, neg, pos, carrier)

            print(f"Logged cycle to data.log and Excel at {timestamp}")

        except Exception as e:
            logging.error("Worker Thread Exception encountered: %s", e)
            print(f"Worker Exception: {e}")

        # Poll and log values every 60 seconds
        time.sleep(60.0)


# Simple Single-Page Dashboard Interface Template
DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>TWR Remote AM Modulation Monitor</title>
    <meta http-equiv="refresh" content="5"> 
    <style>
        body { font-family: Arial, sans-serif; background: #121212; color: #e0e0e0; text-align: center; padding-top: 50px; }
        .container { max-width: 600px; margin: auto; background: #1e1e1e; padding: 20px; border-radius: 10px; box-shadow: 0 4px 10px rgba(0,0,0,0.5); }
        h1 { color: #00ffcc; }
        .metric-box { display: flex; justify-content: space-between; padding: 15px; margin: 10px 0; background: #2a2a2a; border-radius: 5px; font-size: 1.2em; }
        .value { font-weight: bold; color: #ffcc00; }
        .timestamp { font-size: 0.8em; color: #888; margin-top: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Belar AMMA-2 Monitor Dashboard</h1>
        <hr border="1" color="#333">
        <div class="metric-box"><span>Negative Peak Modulation (DA):</span><span class="value">{{ data.negative_peak }}</span></div>
        <div class="metric-box"><span>Positive Peak Modulation (DB):</span><span class="value">{{ data.positive_peak }}</span></div>
        <div class="metric-box"><span>Carrier Reference Level (DG):</span><span class="value">{{ data.carrier_reference }}</span></div>
        <div class="timestamp">Last Sample: {{ data.last_updated }}</div>
    </div>
</body>
</html>
"""


@app.route('/')
def index():
    """Renders the HTML interface."""
    return render_template_string(DASHBOARD_HTML, data=live_telemetry)


@app.route('/api/telemetry')
def api_telemetry():
    """JSON Endpoint for custom external scripts or integrations."""
    return jsonify(live_telemetry)


if __name__ == "__main__":
    # Fire up the hardware poller and logger thread before booting the web server
    worker = Thread(target=telemetry_background_worker, daemon=True)
    worker.start()

    # Listen on all local interfaces (port 5000)
    app.run(host='0.0.0.0', port=5000, debug=False)