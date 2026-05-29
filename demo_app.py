#!/usr/bin/python
"""
TWR Africa — Belar AMMA-2 Real-Life Simulation & Operations Demo

This script simulates real-time AM broadcast audio telemetry feeding into
the thread-safe logging and spreadsheet engines, serving the TWR stylized dashboard.
No physical hardware connection is required for this runtime demonstration.
"""

import logging
import os
import random
from datetime import datetime
from threading import Thread
import time
from flask import Flask, jsonify, render_template_string
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Configure global system file logging
logging.basicConfig(
    filename="./data.log",
    level=logging.INFO,
    format="%(asctime)s : %(message)s",
    datefmt="%Y-%m-%d %H:%M",
)

app = Flask(__name__)

# Shared global telemetry state
live_telemetry = {
    "DA": "0.0 %",
    "DB": "0.0 %",
    "DG": "100.0 %",
    "DH": "-68.5 dB",
    "DN": "0.0 %",
    "last_updated": "Never"
}

EXCEL_FILE = "./modulation_data.xlsx"


def initialize_excel_workbook():
    """Initializes a stylized TWR production ledger if one does not exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modulation Telemetry Log"

        # Apply frozen row scrolling panes
        ws.freeze_panes = "A2"

        headers = [
            "Timestamp Cluster",
            "Negative Peak (DA)",
            "Positive Peak (DB)",
            "Carrier Ref (DG)",
            "AM Noise Floor (DH)",
            "Average RMS (DN)"
        ]

        # Style layout mapping mirroring official TWR specs
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F2C59", end_color="0F2C59", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        wb.save(EXCEL_FILE)


def append_to_excel(data_dict):
    """Appends live runtime data snapshots straight to the spreadsheet archive."""
    try:
        initialize_excel_workbook()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        row_data = [
            data_dict["last_updated"],
            data_dict["DA"],
            data_dict["DB"],
            data_dict["DG"],
            data_dict["DH"],
            data_dict["DN"]
        ]
        ws.append(row_data)

        # Dynamic structural cell Auto-Fit configuration block
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[EXCEL ERROR] Unable to commit row tracking entries: {e}")


def broadcast_simulation_worker():
    """
    Simulates on-air program audio telemetry.
    Generates dynamic audio peak swings and steady carrier parameters.
    """
    print("[SYSTEM ENGINE] Initializing real-life broadcast data telemetry pipeline...")

    while True:
        # 1. Simulate active studio audio program modulation behavior
        # Average volume level swings between 45% and 75% depth
        base_audio_density = random.uniform(45.0, 75.0)

        # Audio spikes create instantaneous peak deviations
        sim_dn = base_audio_density
        sim_da = min(sim_dn + random.uniform(15.0, 35.0), 100.0)  # Cap negative peak safely at 100%
        sim_db = sim_dn + random.uniform(15.0, 48.0)  # Positive peaks can cross 100% b can be greater than or equal to 100% but a cannot exceed 100%

        # Transmitter baseline RF carrier levels stay steady near 100% calibration
        sim_dg = random.uniform(99.8, 100.2)

        # High-performance transmitter noise floor stays down near -70 dB to -65 dB
        sim_dh = random.uniform(-69.2, -66.1)

        # 2. Format variables matching native Belar hardware string formats
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        live_telemetry["DA"] = f"{sim_da:.1f} %"
        live_telemetry["DB"] = f"{sim_db:.1f} %"
        live_telemetry["DG"] = f"{sim_dg:.1f} %"
        live_telemetry["DH"] = f"{sim_dh:.1f} dB"
        live_telemetry["DN"] = f"{sim_dn:.1f} %"
        live_telemetry["last_updated"] = current_time

        # 3. Simulate sequential physical polling execution delays (Writing text log)
        for cmd in ["DA", "DB", "DG", "DH", "DN"]:
            raw_sim_line = f";Data for {cmd} :; {live_telemetry[cmd]}"
            logging.info(raw_sim_line)
            time.sleep(0.1)  # 100ms hardware response latency simulator

        # 4. Save snapshots to the professional corporate spreadsheet ledger
        append_to_excel(live_telemetry)

        print(
            f"[{current_time}] Telemetry snapshot saved. Peak Modulation: +{live_telemetry['DB']} / -{live_telemetry['DA']}")

        # Sleep for a 5-second interval cycle before refreshing fields again
        time.sleep(5)


# TWR Africa Web UI Embedded Template
DASHBOARD_HTML: str = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TWR Africa — Belar AMMA-2 Real-Time Modulation Monitor</title>
    <style>
        :root {
            --bg-main: #0b0f19;
            --bg-card: #111827;
            --border-color: #1f2937;
            --text-main: #f3f4f6;
            --text-muted: #9ca3af;
            --twr-blue: #00a3e0;
            --twr-navy: #0f2c59;
            --signal-ok: #10b981;
            --signal-warn: #f59e0b;
            --signal-crit: #ef4444;
            --led-off: #1f2937;
        }

        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, sans-serif;
            background-color: var(--bg-main);
            color: var(--text-main);
            min-height: 100vh;
            padding: 24px;
        }

        .container {
            max-width: 1000px;
            margin: 0 auto;
        }

        /* Broadcast Console Rack-Mount Header Layout */
        .header {
            background: linear-gradient(90deg, var(--twr-navy) 0%, #071630 100%);
            padding: 20px 24px;
            border-radius: 6px 6px 0 0;
            border-bottom: 3px solid var(--twr-blue);
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 16px;
        }

        .header-title h1 {
            font-size: 1.4rem;
            font-weight: 700;
            letter-spacing: 0.5px;
            color: #ffffff;
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .header-title h2 {
            color: var(--twr-blue);
            font-size: 0.85rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-top: 2px;
        }

        .system-badge {
            background: rgba(0, 163, 224, 0.15);
            border: 1px solid var(--twr-blue);
            color: #ffffff;
            font-size: 0.7rem;
            padding: 4px 10px;
            border-radius: 4px;
            font-weight: 700;
            letter-spacing: 0.5px;
        }

        /* Master Control Status Bar */
        .status-bar {
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-top: none;
            padding: 12px 24px;
            border-radius: 0 0 6px 6px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            font-size: 0.85rem;
        }

        .status-indicator {
            display: flex;
            align-items: center;
            gap: 10px;
            font-weight: 600;
            letter-spacing: 0.5px;
        }

        .pulse-dot {
            width: 10px;
            height: 10px;
            background-color: var(--signal-ok);
            border-radius: 50%;
            animation: telemetryPulse 1.5s ease-in-out infinite;
            box-shadow: 0 0 8px var(--signal-ok);
        }

        @keyframes telemetryPulse {
            0%, 100% { opacity: 1; transform: scale(1); }
            50% { opacity: 0.4; transform: scale(1.1); }
        }

        .refresh-controls {
            display: flex;
            align-items: center;
            gap: 16px;
            color: var(--text-muted);
        }

        .refresh-btn {
            background: #1f2937;
            border: 1px solid #374151;
            color: var(--text-main);
            padding: 4px 12px;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s ease;
        }

        .refresh-btn:hover {
            background: #374151;
            border-color: #4b5563;
        }

        /* High-Density Signal Monitoring Grid Layout */
        .telemetry-grid {
            display: flex;
            flex-direction: column;
            gap: 12px;
            margin-bottom: 20px;
        }

        .telemetry-card {
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            padding: 16px 20px;
            display: grid;
            grid-template-columns: 280px 1fr 140px;
            align-items: center;
            gap: 24px;
            position: relative;
        }

        /* Layout fallback adjustments for ancillary metrics without hardware meters */
        .telemetry-card.no-meter {
            grid-template-columns: 280px 1fr 140px;
        }

        .telemetry-card.no-meter .meter-wrapper {
            display: flex;
            align-items: center;
            color: #4b5563;
            font-size: 0.8rem;
            font-style: italic;
            letter-spacing: 0.5px;
        }

        .metric-details {
            display: flex;
            flex-direction: column;
            gap: 4px;
        }

        .metric-tag {
            font-size: 0.85rem;
            font-weight: 700;
            color: #ffffff;
            letter-spacing: 0.5px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .metric-tag span {
            font-size: 0.7rem;
            background: #1f2937;
            padding: 1px 6px;
            border-radius: 3px;
            color: var(--twr-blue);
        }

        .metric-desc {
            font-size: 0.75rem;
            color: var(--text-muted);
        }

        /* Hardware LED Segment Display Engineering Emulator */
        .meter-wrapper {
            width: 100%;
        }

        .led-strip {
            display: flex;
            gap: 3px;
            background: #020617;
            padding: 6px;
            border-radius: 4px;
            border: 1px solid #1e293b;
        }

        .led-segment {
            flex: 1;
            height: 14px;
            background-color: var(--led-off);
            border-radius: 1px;
            transition: background-color 0.1s ease;
        }

        /* Master Scale Labels below the meter segments */
        .scale-labels {
            display: flex;
            justify-content: space-between;
            margin-top: 4px;
            font-family: 'Courier New', monospace;
            font-size: 0.65rem;
            color: #4b5563;
            font-weight: 700;
            padding: 0 2px;
        }

        /* Digital Nixie Tube Display Box */
        .readout-container {
            text-align: right;
        }

        .digital-readout {
            font-family: 'Courier New', monospace;
            font-size: 1.7rem;
            font-weight: 700;
            color: var(--signal-ok);
            background-color: #050b14;
            border: 1px solid #1e293b;
            padding: 6px 12px;
            border-radius: 4px;
            display: inline-block;
            min-width: 120px;
            text-align: center;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.8);
            letter-spacing: -0.5px;
        }

        /* Specific Status Variations based on RF compliance math thresholds */
        .telemetry-card.alert-flash .digital-readout {
            color: var(--signal-crit) !important;
            animation: criticalFlash 0.5s ease-in-out infinite alternate;
        }

        @keyframes criticalFlash {
            0% { box-shadow: inset 0 2px 4px rgba(0,0,0,0.8); }
            100% { box-shadow: inset 0 0 12px rgba(239, 68, 68, 0.4); border-color: var(--signal-crit); }
        }

        .telemetry-card.warn-state .digital-readout {
            color: var(--signal-warn) !important;
        }

        /* Monitoring Station System Footer Configuration */
        .footer {
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 6px;
            padding: 16px 24px;
            font-size: 0.8rem;
        }

        .sync-time {
            display: flex;
            align-items: center;
            gap: 8px;
            margin-bottom: 12px;
            color: var(--text-main);
        }

        .sync-time strong {
            color: var(--twr-blue);
            font-family: 'Courier New', monospace;
        }

        .metadata-line {
            display: flex;
            gap: 24px;
            flex-wrap: wrap;
            border-top: 1px solid var(--border-color);
            padding-top: 12px;
            color: var(--text-muted);
        }

        .metadata-line span strong {
            color: var(--text-main);
        }

        /* Responsive Breakpoints for Field Inspection Terminals */
        @media (max-width: 850px) {
            .telemetry-card {
                grid-template-columns: 1fr;
                gap: 12px;
            }
            .readout-container {
                text-align: left;
            }
            .digital-readout {
                width: 100%;
            }
        }

        /* Live Data Transit Interception Flash */
        .data-pulse {
            animation: visualCommit 0.3s ease-out;
        }

        @keyframes visualCommit {
            0% { opacity: 0.7; transform: scale(0.995); }
            100% { opacity: 1; transform: scale(1); }
        }
    </style>
</head>
<body>
    <div class="container">
        <!-- Rack Frame Top Header -->
        <div class="header">
            <div class="header-title">
                <h1>📡 TWR AFRICA — AM MONITORING BRIDGE</h1>
                <h2>Belar AMMA-2 Precision Receiver Node</h2>
            </div>
            <div class="system-badge">LINK: TELEMETRY DOWNLINK</div>
        </div>

        <!-- System Diagnostic Line -->
        <div class="status-bar">
            <div class="status-indicator">
                <div class="pulse-dot" id="sysPulse"></div>
                <span id="sysStatusText">CONSOLE STREAMING ACTIVE</span>
            </div>
            <div class="refresh-controls">
                <span>CADENCE: <strong id="refreshCounter" style="color: var(--text-main);">5</strong>s</span>
                <button class="refresh-btn" onclick="toggleAutoRefresh()">⏸ PAUSE BUS</button>
            </div>
        </div>

        <!-- Operational Instruments Matrix Stack -->
        <div class="telemetry-grid">

            <!-- INSTRUMENT 1: NEGATIVE PEAK (DA) -->
            <div class="telemetry-card" id="cardDA">
                <div class="metric-details">
                    <div class="metric-tag">NEGATIVE PEAK MODULATION <span>DA</span></div>
                    <div class="metric-desc">Carrier dropout evaluation (Limit: 100% standard envelope safety)</div>
                </div>
                <div class="meter-wrapper">
                    <div class="led-strip" id="stripDA"></div>
                    <div class="scale-labels">
                        <span>0%</span><span>20</span><span>40</span><span>60</span><span>80</span><span>100%</span><span>120</span><span>140%</span>
                    </div>
                </div>
                <div class="readout-container">
                    <div class="digital-readout" id="valueDA">{{ telemetry.DA }}</div>
                </div>
            </div>

            <!-- INSTRUMENT 2: POSITIVE PEAK (DB) -->
            <div class="telemetry-card" id="cardDB">
                <div class="metric-details">
                    <div class="metric-tag">POSITIVE PEAK MODULATION <span>DB</span></div>
                    <div class="metric-desc">Asymmetric envelope modulation crest factor (Max allocation: 125%)</div>
                </div>
                <div class="meter-wrapper">
                    <div class="led-strip" id="stripDB"></div>
                    <div class="scale-labels">
                        <span>0%</span><span>20</span><span>40</span><span>60</span><span>80</span><span>100%</span><span>120</span><span>140%</span>
                    </div>
                </div>
                <div class="readout-container">
                    <div class="digital-readout" id="valueDB">{{ telemetry.DB }}</div>
                </div>
            </div>

            <!-- INSTRUMENT 3: CARRIER REFERENCE (DG) -->
            <div class="telemetry-card no-meter" id="cardDG">
                <div class="metric-details">
                    <div class="metric-tag">UNMODULATED CARRIER REFERENCE <span>DG</span></div>
                    <div class="metric-desc">RF power engine normalization calibration level</div>
                </div>
                <div class="meter-wrapper">
                    <div>[ Dynamic Linear Reference Signal - Static Variable Tracking ]</div>
                </div>
                <div class="readout-container">
                    <div class="digital-readout" id="valueDG" style="color: #38bdf8;">{{ telemetry.DG }}</div>
                </div>
            </div>

            <!-- INSTRUMENT 4: AM NOISE FLOOR (DH) -->
            <div class="telemetry-card no-meter" id="cardDH">
                <div class="metric-details">
                    <div class="metric-tag">RESIDUAL AM NOISE FLOOR <span>DH</span></div>
                    <div class="metric-desc">Integrated thermal noise &amp; residual interference (Target &lt; -50dB)</div>
                </div>
                <div class="meter-wrapper">
                    <div>[ Spectrum Isolation Subsystem - Decibel Logarithmic Value ]</div>
                </div>
                <div class="readout-container">
                    <div class="digital-readout" id="valueDH">{{ telemetry.DH }}</div>
                </div>
            </div>

            <!-- INSTRUMENT 5: RMS MODULATION (DN) -->
            <div class="telemetry-card" id="cardDN">
                <div class="metric-details">
                    <div class="metric-tag">AVERAGE RMS MODULATION DEPTH <span>DN</span></div>
                    <div class="metric-desc">Integrated root-mean-square continuous audio energy levels</div>
                </div>
                <div class="meter-wrapper">
                    <div class="led-strip" id="stripDN"></div>
                    <div class="scale-labels">
                        <span>0%</span><span>20</span><span>40</span><span>60</span><span>80</span><span>100%</span><span>120%</span>
                    </div>
                </div>
                <div class="readout-container">
                    <div class="digital-readout" id="valueDN">{{ telemetry.DN }}</div>
                </div>
            </div>

        </div>

        <!-- Infrastructure Meta Blocks -->
        <div class="footer">
            <div class="sync-time">
                ⏱ Master Bus Sync: <strong id="lastUpdated">{{ telemetry.last_updated }}</strong>
            </div>
            <div class="metadata-line">
                <span>STORAGE SYSTEM: <strong>data.log</strong> &amp; <strong>modulation_data.xlsx</strong></span>
                <span>DATA RATE: <strong>9600 bps (8N1)</strong></span>
                <span>BUS TIMEOUT: <strong>5000 ms</strong></span>
            </div>
        </div>
    </div>

    <script>
        let autoRefresh = true;
        let refreshInterval;
        let countdownInterval;
        const SEGMENT_COUNT = 40; // High-density resolution meter segments

        // Generate Hardware-Style LED segments programmatically
        function structuralBuildLeds() {
            const targets = ['stripDA', 'stripDB', 'stripDN'];
            targets.forEach(id => {
                const container = document.getElementById(id);
                if (!container) return;
                container.innerHTML = '';
                for (let i = 0; i < SEGMENT_COUNT; i++) {
                    const seg = document.createElement('div');
                    seg.className = 'led-segment';
                    container.appendChild(seg);
                }
            });
        }

        function extractNumericVolume(valStr) {
            if (!valStr || valStr.includes("No Data") || valStr.includes("Error")) return null;
            const outMatch = valStr.match(/(-?\d+(?:\.\d+)?)/);
            return outMatch ? parseFloat(outMatch[1]) : null;
        }

        // Render standard color-coded broadcast segments (Green -> Orange -> Crimson)
        function driveHardwareMeter(stripId, percentageValue, warnThreshold, critThreshold) {
            const strip = document.getElementById(stripId);
            if (!strip) return;
            const segments = strip.getElementsByClassName('led-segment');
            const activeCount = Math.round((Math.min(Math.max(percentageValue, 0), 150) / 150) * SEGMENT_COUNT);

            for (let i = 0; i < SEGMENT_COUNT; i++) {
                if (i < activeCount) {
                    const currentPct = (i / SEGMENT_COUNT) * 150;
                    if (currentPct >= critThreshold) {
                        segments[i].style.backgroundColor = 'var(--signal-crit)';
                        segments[i].style.boxShadow = '0 0 6px var(--signal-crit)';
                    } else if (currentPct >= warnThreshold) {
                        segments[i].style.backgroundColor = 'var(--signal-warn)';
                        segments[i].style.boxShadow = '0 0 4px var(--signal-warn)';
                    } else {
                        segments[i].style.backgroundColor = 'var(--signal-ok)';
                        segments[i].style.boxShadow = 'none';
                    }
                } else {
                    segments[i].style.backgroundColor = 'var(--led-off)';
                    segments[i].style.boxShadow = 'none';
                }
            }
        }

        function evaluateInstrumentCompliance() {
            // Processing Instrument DA (Negative Peak)
            const daVal = extractNumericVolume(document.getElementById('valueDA').innerText);
            const cardDA = document.getElementById('cardDA');
            if (daVal !== null) {
                driveHardwareMeter('stripDA', daVal, 85, 99); // Critical danger at 100% negative (carrier cut)
                cardDA.className = 'telemetry-card' + (daVal >= 99 ? ' alert-flash' : daVal >= 85 ? ' warn-state' : '');
            }

            // Processing Instrument DB (Positive Peak)
            const dbVal = extractNumericVolume(document.getElementById('valueDB').innerText);
            const cardDB = document.getElementById('cardDB');
            if (dbVal !== null) {
                driveHardwareMeter('stripDB', dbVal, 100, 125); // Asymmetrical peaks headroom standard
                cardDB.className = 'telemetry-card' + (dbVal >= 125 ? ' alert-flash' : dbVal >= 100 ? ' warn-state' : '');
            }

            // Processing Instrument DH (Noise Floor evaluation in Decibels)
            const dhVal = extractNumericVolume(document.getElementById('valueDH').innerText);
            const cardDH = document.getElementById('cardDH');
            if (dhVal !== null) {
                // For decibels, values closer to 0 (e.g. -40dB) represent higher noise than -60dB.
                cardDH.className = 'telemetry-card no-meter' + (dhVal >= -45 ? ' alert-flash' : dhVal >= -50 ? ' warn-state' : '');
                document.getElementById('valueDH').style.color = dhVal >= -45 ? 'var(--signal-crit)' : dhVal >= -50 ? 'var(--signal-warn)' : 'var(--signal-ok)';
            }

            // Processing Instrument DN (RMS Audio Tracking)
            const dnVal = extractNumericVolume(document.getElementById('valueDN').innerText);
            const cardDN = document.getElementById('cardDN');
            if (dnVal !== null) {
                // Adjusting scale calculation matrix to target 125% volume mapping reference
                const normalizedDN = (dnVal / 125) * 150;
                driveHardwareMeter('stripDN', normalizedDN, 85, 110);
                cardDN.className = 'telemetry-card' + (dnVal >= 110 ? ' alert-flash' : dnVal >= 85 ? ' warn-state' : '');
            }
        }

        function updateDashboard() {
            fetch('/api/telemetry')
                .then(res => res.json())
                .then(data => {
                    document.getElementById('sysPulse').style.backgroundColor = 'var(--signal-ok)';
                    document.getElementById('sysPulse').style.boxShadow = '0 0 8px var(--signal-ok)';
                    document.getElementById('sysStatusText').innerText = "CONSOLE STREAMING ACTIVE";

                    const metrics = ['DA', 'DB', 'DG', 'DH', 'DN'];
                    metrics.forEach(m => {
                        const cell = document.getElementById(`value${m}`);
                        const incoming = data[m] || "No Data";
                        if (cell.innerText !== incoming) {
                            cell.innerText = incoming;
                            const card = document.getElementById(`card${m}`);
                            card.classList.add('data-pulse');
                            setTimeout(() => card.classList.remove('data-pulse'), 300);
                        }
                    });

                    document.getElementById('lastUpdated').innerText = data.last_updated || "Never";
                    evaluateInstrumentCompliance();
                })
                .catch(err => {
                    console.error('Bus Downlink Exception:', err);
                    document.getElementById('sysPulse').style.backgroundColor = 'var(--signal-crit)';
                    document.getElementById('sysPulse').style.boxShadow = '0 0 8px var(--signal-crit)';
                    document.getElementById('sysStatusText').innerText = "LINK TIMEOUT / HARDWARE FAULT";
                });
        }

        function startAutoRefresh() {
            if (refreshInterval) clearInterval(refreshInterval);
            if (countdownInterval) clearInterval(countdownInterval);

            let sec = 5;
            const counter = document.getElementById('refreshCounter');
            if (counter) counter.innerText = sec;

            refreshInterval = setInterval(() => {
                if (autoRefresh) {
                    updateDashboard();
                    sec = 5;
                    if (counter) counter.innerText = sec;
                }
            }, 5000);

            countdownInterval = setInterval(() => {
                if (autoRefresh) {
                    sec--;
                    if (counter && sec >= 0) counter.innerText = sec;
                    if (sec < 0) sec = 4;
                }
            }, 1000);
        }

        function toggleAutoRefresh() {
            autoRefresh = !autoRefresh;
            const btn = document.querySelector('.refresh-btn');
            if (autoRefresh) {
                btn.innerHTML = '⏸ PAUSE BUS';
                btn.style.background = '#1f2937';
                updateDashboard();
                startAutoRefresh();
            } else {
                btn.innerHTML = '▶ RESUME BUS';
                btn.style.background = 'var(--twr-navy)';
                if (refreshInterval) clearInterval(refreshInterval);
                if (countdownInterval) clearInterval(countdownInterval);
            }
        }

        window.addEventListener('load', () => {
            structuralBuildLeds();
            evaluateInstrumentCompliance();
            startAutoRefresh();
        });
    </script>
</body>
</html>
"""
@app.route('/')
def index():
    return render_template_string(DASHBOARD_HTML, telemetry=live_telemetry)


@app.route('/api/telemetry')
def api_telemetry():
    return jsonify(live_telemetry)


if __name__ == "__main__":
    # Start the simulation data loop thread
    sim_thread = Thread(target=broadcast_simulation_worker, daemon=True)
    sim_thread.start()

    # Launch local development web server
    app.run(host='0.0.0.0', port=5000, debug=False)